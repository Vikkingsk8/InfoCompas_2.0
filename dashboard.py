from dash import Dash, dcc, html
from dash.dependencies import Input, Output
import plotly.express as px
import pandas as pd
import requests
from openpyxl import Workbook
from io import BytesIO

def create_dash_app(flask_app, routes_pathname_prefix='/dashboard/'):
    dash_app = Dash(__name__, server=flask_app, routes_pathname_prefix=routes_pathname_prefix)
    dash_app.layout = html.Div([
        html.H1("Аналитика чат-бота", style={'textAlign': 'center', 'color': '#ffffff', 'padding': '20px'}),
        dcc.Interval(
            id='interval-component',
            interval=60*1000,
            n_intervals=0
        ),
        html.Div(id='metrics-container', style={'textAlign': 'center', 'margin': '20px', 'backgroundColor': 'rgba(255, 255, 255, 0.7)', 'padding': '10px', 'borderRadius': '10px'}),
        html.Div([
            html.Div([
                dcc.Graph(id='queries-over-time'),
                dcc.Graph(id='queries-by-day'),
            ], style={'width': '50%', 'display': 'inline-block'}),
            html.Div([
                dcc.Graph(id='queries-by-week'),
                dcc.Graph(id='queries-by-month'),
            ], style={'width': '50%', 'display': 'inline-block'}),
        ]),
        html.Div([
            html.Div([
                dcc.Graph(id='top-questions'),
            ], style={'width': '50%', 'display': 'inline-block'}),
            html.Div([
                dcc.Graph(id='feedback-distribution'),
            ], style={'width': '50%', 'display': 'inline-block'}),
        ]),
        html.Button("Экспорт в Excel", id="export-button", n_clicks=0),
        dcc.Download(id="download-excel")
    ], style={
        'backgroundImage': 'url("/static/background.jpg")',
        'backgroundSize': 'cover',
        'backgroundRepeat': 'no-repeat',
        'backgroundAttachment': 'fixed',
        'backgroundPosition': 'center',
        'padding': '20px',
        'minHeight': '100vh'
    })

    @dash_app.callback(
        [Output('metrics-container', 'children'),
         Output('queries-over-time', 'figure'),
         Output('queries-by-day', 'figure'),
         Output('queries-by-week', 'figure'),
         Output('queries-by-month', 'figure'),
         Output('top-questions', 'figure'),
         Output('feedback-distribution', 'figure')],
        Input('interval-component', 'n_intervals')
    )
    def update_metrics(n):
        try:
            response = requests.get('http://176.109.109.61:8080/analytics_data')
            response.raise_for_status()
            data = response.json()
        except requests.RequestException as e:
            print(f"Ошибка при запросе к API: {e}")
            return html.Div("Ошибка при получении данных"), {}, {}, {}, {}, {}, {}

        metrics = [
            html.P(f"Всего запросов: {data['total_queries']}"),
            html.P(f"Успешных запросов: {data['successful_queries']}"),
            html.P(f"Процент успешных запросов: {data['success_rate']:.2f}%"),
        ]
        
        # График запросов по времени
        queries_df = pd.DataFrame(list(data['queries_over_time'].items()), columns=['time', 'count'])
        queries_df['time'] = pd.to_datetime(queries_df['time'])
        queries_fig = px.line(queries_df, x='time', y='count', title="Запросы по времени")
        
        # График запросов по дням
        queries_by_day = pd.DataFrame(list(data['queries_by_day'].items()), columns=['day', 'count'])
        queries_by_day['day'] = pd.to_datetime(queries_by_day['day'])
        queries_by_day_fig = px.bar(queries_by_day, x='day', y='count', title="Запросы по дням")
        
        # График запросов по неделям
        queries_by_week = pd.DataFrame(list(data['queries_by_week'].items()), columns=['week', 'count'])
        queries_by_week_fig = px.bar(queries_by_week, x='week', y='count', title="Запросы по неделям")
        
        # График запросов по месяцам
        queries_by_month = pd.DataFrame(list(data['queries_by_month'].items()), columns=['month', 'count'])
        queries_by_month_fig = px.bar(queries_by_month, x='month', y='count', title="Запросы по месяцам")
        
        # График топ-10 вопросов
        top_questions = pd.DataFrame(data['top_questions'], columns=['question', 'count'])
        top_questions = top_questions.sort_values('count', ascending=True).tail(10)
        top_questions_fig = px.bar(top_questions, x='count', y='question', orientation='h', title="Топ-10 вопросов")
        
        # График распределения обратной связи
        feedback_df = pd.DataFrame(list(data['feedback_distribution'].items()), columns=['feedback', 'count'])
        feedback_fig = px.pie(feedback_df, values='count', names='feedback', title="Распределение обратной связи")
        
        return metrics, queries_fig, queries_by_day_fig, queries_by_week_fig, queries_by_month_fig, top_questions_fig, feedback_fig

    @dash_app.callback(
        Output("download-excel", "data"),
        Input("export-button", "n_clicks"),
        prevent_initial_call=True
    )
    def export_to_excel(n_clicks):
        if n_clicks == 0:
            return None
    
        try:
            response = requests.get('http://176.109.109.61:8080/analytics_data')
            response.raise_for_status()
            data = response.json()
        except requests.RequestException as e:
            print(f"Ошибка при запросе к API: {e}")
            return None

        wb = Workbook()
    
        # Создаем листы и заполняем их данными
        ws = wb.active
        ws.title = "Общая статистика"
        ws.append(["Метрика", "Значение"])
        ws.append(["Всего запросов", data['total_queries']])
        ws.append(["Успешных запросов", data['successful_queries']])
        ws.append(["Процент успешных запросов", f"{data['success_rate']:.2f}%"])

        for sheet_name, sheet_data in [
            ("Запросы по времени", data['queries_over_time']),
            ("Запросы по дням", data['queries_by_day']),
            ("Запросы по неделям", data['queries_by_week']),
            ("Запросы по месяцам", data['queries_by_month']),
        ]:
            ws = wb.create_sheet(sheet_name)
            ws.append(["Время", "Количество"])
            for time, count in sheet_data.items():
                ws.append([time, count])

        ws = wb.create_sheet("Топ вопросов")
        ws.append(["Вопрос", "Количество"])
        for question in data['top_questions']:
            ws.append(question)

        ws = wb.create_sheet("Обратная связь")
        ws.append(["Оценка", "Количество"])
        for feedback, count in data['feedback_distribution'].items():
            ws.append([feedback, count])

        # Сохраняем файл в памяти
        excel_buffer = BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        return dcc.send_bytes(excel_buffer.getvalue(), "dashboard_data.xlsx")

    return dash_app